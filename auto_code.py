import time
import threading
from openpyxl import load_workbook
from selenium import webdriver
import keyboard
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import tkinter as tk
import collections
from tkinter import filedialog
from selenium.common.exceptions import TimeoutException
import tkinter.messagebox as messagebox
import os
import sys
import csv
import re
import pandas as pd
from fake_useragent import UserAgent

continue_event = threading.Event()


def read_excel_file(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    return sheet


def read_first_column_csv(file_path):
    first_column = {}
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            if row:  # To handle empty rows
                cleaned_text = re.sub(r'[^a-zA-Z0-9]+', '', row[0]).upper()
                first_column[cleaned_text] = True

    return first_column


def read_first_column_excel(file_path):
    first_column = {}
    df = pd.read_excel(file_path)
    column_values = df.iloc[:, 0].dropna().values

    for value in column_values:
        cleaned_text = re.sub(r'[^a-zA-Z0-9]+', '', str(value)).upper()
        first_column[cleaned_text] = True

    return first_column


def on_continue_click():
    continue_event.set()


def read_csv_file(file_path):
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        return list(reader)


def write_message_to_file(message):
    with open('code_status.txt', 'w') as f:
        f.write(message)


def show_popup_message(merged_error_codes):
    message = ""

    if len(merged_error_codes) == 0:
        message = "\nNo issues found."
    else:
        message += "\nIssues found. More info checked the code_status file, located in the same directory of the program."
        if len(merged_error_codes["Invalid Code"]) > 0:
            message += "\nThe following codes are invalid:\n"
            message += "\n".join(merged_error_codes["Invalid Code"])

        if len(merged_error_codes["Redeemed"]) > 0:
            message += "\nThe following codes have already been redeemed:\n"
            message += "\n".join(merged_error_codes["Redeemed"])

        if len(merged_error_codes["Miss"]) > 0:
            message += "\nThe following codes are missing from automation. Please add manually:\n"
            message += "\n".join(merged_error_codes["Miss"])

    write_message_to_file(message)
    messagebox.showinfo("Code Status", message)


def on_close(root):
    if messagebox.askokcancel("Quit", "Do you really want to quit?"):
        root.destroy()


def start_main_thread(initial_row, browser, copy_count, full_automation, start_button, continue_button, sleep_time_var):
    # user click start：Disable the button and change its color to grey
    if not file_path_var.get():
        messagebox.showwarning(
            "Warning", "No files loaded. Please select a file before starting.")
        return
    start_button.config(state="disabled")
    sleep_time = float(sleep_time_var.get())
    # if in full automation mode, change continue button to grey so that users can not click
    if full_automation.get():
        continue_button.config(state="disabled")
    # Create and start the thread
    thread = threading.Thread(target=run_main_thread, args=(
        initial_row, browser, copy_count, full_automation, sleep_time))
    thread.start()

    # Check the thread's status and update the button state accordingly
    check_thread_status(start_button, thread)


def check_thread_status(start_button, thread):
    if thread.is_alive():
        # If the thread is still running, check its status again after 100 milliseconds
        start_button.after(
            100, lambda: check_thread_status(start_button, thread))
    else:
        # If the thread has finished, re-enable the button and reset its color
        start_button.config(state="normal", bg="SystemButtonFace")


def run_main_thread(initial_row, browser, copy_count, full_automation, sleep_time):
    # redeemed_codes = set()
    merged_error_codes = collections.defaultdict(set)
    more_codes = True
    file_path = file_path_var.get()
    remainingCodes = read_first_column_excel(file_path)
    if full_automation.get():
        while True:
            initial_row, copied_codes, more_codes, error_codes = main(
                initial_row, browser, copy_count, full_automation, sleep_time, remainingCodes)
            for key in set(merged_error_codes.keys()) | set(error_codes.keys()):
                merged_error_codes[key] = merged_error_codes[key] | error_codes[key]
            if not more_codes:
                break
    else:
        first_group = True
        while more_codes:
            if not first_group:
                # Wait for Ctrl+Space before copying each group of 10 codes
                # keyboard.wait("ctrl+space")
                continue_event.wait()
                continue_event.clear()

            initial_row, copied_codes, more_codes, error_codes = main(
                initial_row, browser, 10, full_automation, sleep_time, remainingCodes)
            for key in set(merged_error_codes.keys()) | set(error_codes.keys()):
                merged_error_codes[key] = merged_error_codes[key] | error_codes[key]

            if not more_codes:
                break

            first_group = False
    # handle missed codes
    for code in remainingCodes:
        if remainingCodes[code] == True:
            merged_error_codes["Miss"].add(code)
    show_popup_message(merged_error_codes)
    if len(merged_error_codes) == 0:
        print("\n No issuse found.")

    if len(merged_error_codes["Invalid Code"]) > 0:
        print("\nThe following codes are invalid:")
        for code in merged_error_codes["Invalid Code"]:
            print(code)
    if len(merged_error_codes["Redeemed"]) > 0:
        print("\nThe following codes have already been redeemed:")
        for code in merged_error_codes["Redeemed"]:
            print(code)
    if len(merged_error_codes["Miss"]) > 0:
        print("\nThe following codes are missing from automation. Please add manually:")
        for code in merged_error_codes["Miss"]:
            print(code)


def select_file():
    file_path = filedialog.askopenfilename(defaultextension=".xlsx")
    if file_path:
        file_path_var.set(file_path)


def start_app():
    root = tk.Tk()
    root.title("autoRedeem-WebVersion")
    root.protocol("WM_DELETE_WINDOW", lambda: on_close(root))

    if getattr(sys, 'frozen', False):
        script_dir = sys._MEIPASS  # Use the bundled app directory if running in the executable
    else:
        # Use the script directory if running in the Python script
        script_dir = os.path.dirname(os.path.realpath(__file__))

    icon_path = os.path.join(script_dir, 'Mimikyu.ico')
    root.iconbitmap(icon_path)

    global file_path_var
    file_path_var = tk.StringVar()
    file_path_var.set("")

    tk.Label(root, text="Excel File Path:").grid(row=0, column=0, sticky="e")
    tk.Entry(root, textvariable=file_path_var,
             width=50).grid(row=0, column=1, padx=5)
    tk.Button(root, text="Browse", command=select_file).grid(
        row=0, column=2, padx=5)

    full_automation = tk.BooleanVar()
    tk.Checkbutton(root, text="Full Automation",
                   variable=full_automation).grid(row=2, column=1, pady=5)

    continue_var = tk.BooleanVar()
    continue_var.set(False)

    continue_button = tk.Button(
        root, text="Continue", command=on_continue_click)
    continue_button.grid(row=5, column=1, pady=10)

    sleep_time_var = tk.StringVar()
    sleep_time_var.set("1")
    tk.Label(root, text="Sleep Time:").grid(row=1, column=0, sticky="w")
    tk.Entry(root, textvariable=sleep_time_var).grid(
        row=1, column=1, padx=3, sticky="w")

    start_button = tk.Button(root, text="Start Script", command=lambda: start_main_thread(
        initial_row, browser, copy_count, full_automation, start_button, continue_button, sleep_time_var))

    start_button.grid(row=3, column=1, pady=10)

    if not full_automation:
        continue_button.config(state="disabled")

    author_label = tk.Label(root, text="版权所有: @闲鱼: 巨糕冷")
    author_label.grid(row=6, column=1, pady=10)
    root.mainloop()


def main(initial_row, browser, copy_count, full_automation, sleep_time, remainingCodes):
    file_path = file_path_var.get()
    sheet = read_excel_file(file_path)
    # sheet = read_csv_file(file_path)
    #remainingCodes = read_first_column_csv(file_path)
    copied_codes = 0
    more_codes = True
    # remainingCodes = read_first_column_excel(file_path)
    new_redeemed_codes = set()
    error = collections.defaultdict(set)

    should_clear_table = False

    while (full_automation.get() and more_codes) or (copied_codes < copy_count and more_codes):
        #cell_value = sheet[initial_row][0]
        cell_value = sheet.cell(row=initial_row, column=1).value
        if not cell_value:
            print("You have copied all the codes.")
            more_codes = False
            should_clear_table = True
            break

        try:
            input_box = WebDriverWait(browser, 10).until(
                EC.presence_of_element_located((By.ID, "code")))
            submit_button = WebDriverWait(browser, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'button[data-testid="verify-code-button"]')))

            input_box.send_keys(cell_value)
            submit_button.click()

            time.sleep(sleep_time)
            initial_row += 1
            copied_codes += 1
            lastOne = False
            if full_automation.get():
                try:
                    # next_cell_value = sheet[initial_row][0]
                    next_cell_value = sheet.cell(
                        row=initial_row, column=1).value
                    if not next_cell_value:
                        lastOne = True
                    # if (copied_codes % 10 == 0) or (not next_cell_value):
                    if (copied_codes % 10 == 0):
                        # for self test
                        # redeem_button = WebDriverWait(browser, 10).until(EC.presence_of_element_located(
                        #     (By.CSS_SELECTOR, 'button[data-testid="button-redeem"]')))
                        # redeem_button.click()
                        time.sleep(3)
                        clear_table_button = WebDriverWait(browser, 10).until(EC.presence_of_element_located(
                            (By.CSS_SELECTOR, 'button[data-testid="button-clear-table"]')))
                        clear_table_button.click()

                except TimeoutException:
                    print("Timed out waiting for the clear table/redeem button.")

        except TimeoutException:
            print("Timed out waiting for the elements to load.")
            break

        # Check the status of the codes after copying
        code_elements = browser.find_elements(
            By.CSS_SELECTOR, 'td.RedeemModule_tdCode__2V387')
        status_elements = browser.find_elements(
            By.CSS_SELECTOR, 'td.RedeemModule_tdLocalizedName__1VWAC')

        for code_element, status_element in zip(code_elements, status_elements):
            status_text = status_element.text.strip()
            # print(code_element.text.strip())
            redeemed_code = code_element.text.strip()
            clean_redeemed_code = re.sub(
                r'[^a-zA-Z0-9]+', '', redeemed_code).upper()
            if "This code has already been redeemed" in status_text:
                error["Redeemed"].add(redeemed_code)
            if "Invalid Code" in status_text:
                error["Invalid Code"].add(redeemed_code)
            remainingCodes[clean_redeemed_code] = False
            # brutally test missing code
            # if clean_redeemed_code == '9TLKVRXZBDHNV':
            #     remainingCodes[clean_redeemed_code] = True
            print('===================')
            print(clean_redeemed_code)
            print(remainingCodes)
            print(code_element)
            print(status_element)

        if lastOne == True:
            # redeem_button = WebDriverWait(browser, 10).until(EC.presence_of_element_located(
            #     (By.CSS_SELECTOR, 'button[data-testid="button-redeem"]')))
            # redeem_button.click()
            time.sleep(2)
            clear_table_button = WebDriverWait(browser, 10).until(EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'button[data-testid="button-clear-table"]')))
            clear_table_button.click()
    return initial_row, copied_codes, more_codes, error


if __name__ == "__main__":
    # initial_row = 0
    initial_row = 1
    loop_finished = False
    copy_count = 10

    chrome_options = Options()
    ua = UserAgent()
    user_agent = ua.random

    # Add the user agent argument to Chrome options
    chrome_options.add_argument(f'user-agent={user_agent}')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    # chrome_options.add_argument(
    #     '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36')

    browser = webdriver.Chrome(options=chrome_options)

    url = "https://redeem.tcg.pokemon.com/en-us/"
    browser.get(url)

    print("Please log in manually. The script will continue after you have logged in.")
    WebDriverWait(browser, 300).until(
        EC.presence_of_element_located((By.ID, "code")))

    start_app()
    browser.quit()
    print("Exiting...")
